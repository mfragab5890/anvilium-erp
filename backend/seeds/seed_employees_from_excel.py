"""
Seed employees from an Excel file (salary multiplied by 0.317).

Run from backend folder:
  (.venv) python -m seeds.seed_employees_from_excel
or:
  (.venv) python seeds\seed_employees_from_excel.py
"""
from __future__ import annotations
import argparse, os, math
from datetime import datetime
from typing import List, Dict, Any, Optional

from app import create_app
from extensions import db
import sqlalchemy as sa

from modules.core.models import Branch
from modules.hr.models import Employee

DEFAULT_XLSX_PATH = r"D:\Web Dev\mine\ANVILIUM\21- ANVILIUM DRIVE\2- ANVILIUM LABORS\1- Workmen list\Labor List_2025.XLSX"

# ---------- Excel reading ----------
def _try_import_excel():
    try:
        import pandas as pd  # type: ignore
        return ("pandas", pd)
    except Exception:
        try:
            from openpyxl import load_workbook  # type: ignore
            return ("openpyxl", load_workbook)
        except Exception:
            return (None, None)

def _read_rows_xlsx(path: str) -> List[Dict[str, Any]]:
    mode, lib = _try_import_excel()
    if mode == "pandas":
        pd = lib
        df = pd.read_excel(path, sheet_name=0)
        df.columns = [str(c).strip().lower() for c in df.columns]
        return df.to_dict(orient="records")
    elif mode == "openpyxl":
        load_workbook = lib
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        data: List[Dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2):
            row: Dict[str, Any] = {}
            for i, c in enumerate(r):
                key = headers[i] if i < len(headers) else f"col{i}"
                row[key] = c.value
            data.append(row)
        return data
    else:
        raise RuntimeError("Please install either pandas or openpyxl to read Excel files.")

# ---------- helpers ----------
def _first_nonempty(d: Dict[str, Any], keys: List[str], default: str = "") -> str:
    for k in keys:
        v = d.get(k)
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return default

def _to_float(v: Any) -> float:
    try:
        if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
            return 0.0
        return float(str(v).strip().replace(",", ""))
    except Exception:
        return 0.0

def _to_date(v: Any):
    # very tolerant date parser; pandas usually gives datetime already
    try:
        import pandas as pd  # type: ignore
        if isinstance(v, (str, bytes)):
            s = str(v).strip()
            if not s:
                return None
            return pd.to_datetime(s, errors="coerce").date()
        elif hasattr(v, "date"):
            return v.date()
        return None
    except Exception:
        return None

# Build one employee dict from a raw Excel row
def transform_row(r: Dict[str, Any], idx: int, branch_by_code: dict[str,int], branch_by_name: dict[str,int]) -> Dict[str, Any]:
    code  = _first_nonempty(r, ["code", "emp_code", "employee code", "employee_code"], f"E{idx:04d}")
    email = _first_nonempty(r, ["email", "mail"], f"emp{idx}@example.com").lower()

    first = _first_nonempty(r, ["first_name", "first", "given name", "name_en", "name"], "")
    last  = _first_nonempty(r, ["last_name", "last", "surname"], "")
    if not first:
        full = _first_nonempty(r, ["employee name", "employee_name"], "")
        if full:
            parts = full.split()
            first = parts[0] if parts else f"Emp{idx}"
            last  = " ".join(parts[1:]) if len(parts) > 1 else ""

    phone = _first_nonempty(r, ["phone", "mobile", "contact", "phone no", "mobile no"], "")
    position = _first_nonempty(r, ["position", "designation", "profession", "trade", "job title"], "")
    nationality = _first_nonempty(r, ["nationality", "country"], "")
    dob = _to_date(_first_nonempty(r, ["dob", "date of birth", "birth date"], ""))

    # salary × 0.317
    salary_raw = _to_float(_first_nonempty(r, ["salary", "monthly_salary", "basic", "base", "wage"], "0"))
    salary_monthly = round(salary_raw * 0.317, 2)

    # branch resolve
    branch_str = _first_nonempty(r, ["branch", "site", "location", "office"], "")
    branch_id: Optional[int] = None
    if branch_str:
        key = branch_str.strip().lower()
        branch_id = branch_by_code.get(key) or branch_by_name.get(key)

    # Collect everything else into meta (raw snapshot)
    known_keys = {
        "code","emp_code","employee code","employee_code",
        "email","mail",
        "first_name","first","given name","name_en","name","employee name","employee_name",
        "last_name","last","surname",
        "phone","mobile","contact","phone no","mobile no",
        "position","designation","profession","trade","job title",
        "nationality","country",
        "dob","date of birth","birth date",
        "salary","monthly_salary","basic","base","wage",
        "branch","site","location", "office",
    }
    meta = {k: v for k, v in r.items() if k not in known_keys}

    out = dict(
        code=code,
        first_name=first or f"Emp{idx}",
        last_name=last or "",
        email=email,
        phone=phone or None,
        position=position or None,
        branch_id=branch_id,
        hire_date=None,              # we will fill when/if we know which header carries it
        termination_date=None,
        is_active=True,
        salary_monthly=salary_monthly,
        nationality=nationality or None,
        dob=dob,
        meta=meta or None,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    return out

# constrain to real table columns (safe across DBs)
def _filter_to_table_columns(row: Dict[str, Any], table: sa.Table) -> Dict[str, Any]:
    cols = {c.name for c in table.c}
    return {k: v for k, v in row.items() if k in cols}

def bulk_insert_employees(rows: List[Dict[str, Any]]) -> int:
    if not rows:
        return 0
    table = sa.inspect(Employee).local_table
    filtered = [_filter_to_table_columns(r, table) for r in rows]
    filtered = [r for r in filtered if r.get("code") or r.get("email")]
    if not filtered:
        return 0

    dialect = db.engine.dialect.name
    conflict_col = "code" if "code" in table.c else ("email" if "email" in table.c else None)

    if dialect == "postgresql":
        from sqlalchemy.dialects.postgresql import insert as psql_insert
        stmt = psql_insert(table).values(filtered)
        if conflict_col:
            stmt = stmt.on_conflict_do_nothing(index_elements=[table.c[conflict_col]])
        else:
            stmt = stmt.on_conflict_do_nothing()
        res = db.session.execute(stmt)
        return res.rowcount or 0
    elif dialect in {"mysql", "mariadb"}:
        from sqlalchemy.dialects.mysql import insert as mysql_insert
        stmt = mysql_insert(table).values(filtered).prefix_with("IGNORE")
        res = db.session.execute(stmt)
        return res.rowcount or 0
    else:
        stmt = sa.insert(table).values(filtered).prefix_with("OR IGNORE")
        res = db.session.execute(stmt)
        return res.rowcount or 0

def main(xlsx_path: str) -> None:
    # Preload branches for quick lookup
    branches = Branch.query.all()
    branch_by_code = { (b.code or "").strip().lower(): b.id for b in branches if b.code }
    branch_by_name = { (b.name or "").strip().lower(): b.id for b in branches if b.name }

    raw_rows = _read_rows_xlsx(xlsx_path)
    prepared = [transform_row(r, i + 1, branch_by_code, branch_by_name) for i, r in enumerate(raw_rows)]
    inserted = bulk_insert_employees(prepared)
    db.session.commit()
    print(f"Employees inserted (duplicates ignored by code/email): {inserted}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", dest="path",
                        default=os.getenv("EMP_XLSX") or os.getenv("EMPLOYEES_XLSX_PATH") or DEFAULT_XLSX_PATH)
    args = parser.parse_args()

    if not args.path or not os.path.exists(args.path):
        raise SystemExit(f"Excel file not found:\n  {args.path}\nUpdate DEFAULT_XLSX_PATH or pass --path.")

    app = create_app()
    with app.app_context():
        main(args.path)
